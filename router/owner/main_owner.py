from datetime import datetime
import os
import stat
from typing import Optional
import uuid
import aiomysql
from fastapi import APIRouter, File, Form, Query, Request, HTTPException, Security, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from koneksi import get_db
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.workbook.protection import WorkbookProtection
from collections import defaultdict
import pandas as pd
from aiomysql import Error as aiomysqlerror
import asyncio
import win32com.client
from pywintypes import com_error

app = APIRouter(prefix=("/main_owner"))

@app.get('/get_laporan')
async def getLaporan() :
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
        # Pilih Awal dlu
        qMain = "SELECT id_transaksi, disc, gtotal_stlh_pajak, total_harga, jenis_pembayaran FROM main_transaksi WHERE status IN ('paid', 'done') AND is_cancel = 0"
        await cursor.execute(qMain)  
        itemMain = await cursor.fetchall()

        # Buat dictionary main_transaksi
        main_trans_dict = {item['id_transaksi']: item for item in itemMain}
        """
          hasil main_trans_dict : 
          {
            "TF0001": {
              "id_transaksi": "TF0001",
              "disc": 0.2,
              "gtotal_stlh_pajak": 648000,
              "total_harga": 810000
            },
            "TF0002": { dan seterusnya},
          }
        """

        """
          Fungsi Query1 Ini : (Utk Line Chart)
          1. Filters records where created_at is from the first day of the current month
          2. Up to (but not including) the first day of the month 12 months from now
          3. Groups by year-month
          4. Orders chronologically
          5. Limits to 12 months of results
        """
        # q1 = """

        #   SELECT DATE_FORMAT(created_at, "%Y-%m") AS bulan,
        #   SUM(gtotal_stlh_pajak) AS omset_jual
        #   FROM main_transaksi
        #   WHERE created_at >= DATE_FORMAT(CURRENT_DATE, '%Y-%m-01') 
        #   AND created_at < DATE_FORMAT(DATE_ADD(CURRENT_DATE, INTERVAL 12 MONTH), '%Y-%m-01')
        #   GROUP BY bulan
        #   LIMIT 12
        # """
        q1 = """
          WITH RECURSIVE months AS (
            SELECT DATE_FORMAT(DATE_SUB(CURRENT_DATE, INTERVAL 11 MONTH), '%Y-%m') AS bulan,
            1 AS num
            UNION ALL
            SELECT DATE_FORMAT(DATE_ADD(DATE_SUB(CURRENT_DATE, INTERVAL 11 MONTH), INTERVAL num MONTH), '%Y-%m'), num + 1
            FROM months 
            WHERE num < 12
          )
          SELECT m.bulan, IFNULL(SUM(t.gtotal_stlh_pajak), 0) AS omset_jual
          FROM months m 
          LEFT JOIN main_transaksi t 
          ON DATE_FORMAT(t.created_at, '%Y-%m') = m.bulan AND t.status IN ('done', 'paid') AND t.is_cancel = 0
          GROUP BY m.bulan
          ORDER BY m.bulan
        """
        await cursor.execute(q1)  
        items1 = await cursor.fetchall()

        # Utk Perbandingan Current Month sama Sebelumnya. Total Penjualan (Paket Produk)
        queryWith = """
          WITH months AS (
            SELECT DATE_FORMAT(DATE_SUB(CURRENT_DATE, INTERVAL 1 MONTH), '%Y-%m') AS bulan
            UNION ALL
            SELECT DATE_FORMAT(CURRENT_DATE, '%Y-%m') AS bulan
          )
        """
        # Utk Dapetin Monthly Sales
        q2 = f"""
          {queryWith}
          SELECT m.bulan, IFNULL(SUM(t.gtotal_stlh_pajak), 0) AS omset_jual
          FROM months m
          LEFT JOIN main_transaksi t 
            ON DATE_FORMAT(t.created_at, '%Y-%m') = m.bulan 
            AND t.status IN ('done', 'paid') AND t.is_cancel = 0

          GROUP BY m.bulan
          ORDER BY m.bulan DESC
        """
        await cursor.execute(q2)
        items2 = await cursor.fetchall()

        # Penjualan Paket. ambil dari gtotal_stlh_pajak yg main_transaksi. 
        q3 = f"""
          {queryWith}
          SELECT
              m.bulan,
              -- 1. Mengganti kolom menjadi gtotal_stlh_pajak
              IFNULL(SUM(t.gtotal_stlh_pajak), 0) AS omset_bulanan
          FROM
              months m
          LEFT JOIN
              main_transaksi t ON DATE_FORMAT(t.created_at, '%Y-%m') = m.bulan
              AND t.status IN ('paid', 'done')
              AND t.is_cancel = 0
              -- 2. Mengganti filter EXISTS dengan jenis_transaksi
              AND t.jenis_transaksi = 'massage'
          GROUP BY
              m.bulan
          ORDER BY
              m.bulan DESC;
        """
        await cursor.execute(q3)
        items3 = await cursor.fetchall()

        # penjualan_paket = []
        # for data in items3:
        #   id_transaksi = data['id_transaksi']
        #   omset_jual_paket = int(data['omset_jual_paket']) or 0

        #   # ambil key id_transaksi di maintrans yg sama dgn data['id_transaksi']
        #   get_same_id = main_trans_dict.get(id_transaksi, {}) #klo g ad return obj kosong
        #   disc = float(get_same_id.get('disc', 0)) #klo g ad maka 0
          
        #   # Rumus Persen
        #   # if data['jenis_pembayaran'] == 1:
        #   #   # Case 1: Bayar diakhir, pasti dpt disc
        #   #   nominal_disc = omset_jual_paket * disc
        #   #   harga_stlh_disc = omset_jual_paket - nominal_disc
        #   # elif data['jenis_pembayaran'] == 0 and data['is_addon'] == 0:
        #   #   # Case 2: Payment upfront - only apply discount to non-addon items
        #   #   nominal_disc = omset_jual_paket * disc
        #   #   harga_stlh_disc = omset_jual_paket - nominal_disc
        #   # else:
        #   #   # Ga dpt disc (bayar diawal. addon g dpt)
        #   #   harga_stlh_disc = omset_jual_paket

        #   # Update key items3
        #   data['omset_jual_paket'] = omset_jual_paket
        #   penjualan_paket.append(data)

        # # GroupBy valuenya berdasarkan key bulan
        # monthly_paket = {}
        # for entry in penjualan_paket:
        #   bulan = entry['bulan']
        #   if bulan in monthly_paket:
        #     monthly_paket[bulan] += entry['omset_jual_paket']
        #   else:
        #     monthly_paket[bulan] = entry['omset_jual_paket']

        # # restruktur ulang dictionarynya, lalujadikan dataframe biar bentuk kek tabel sql
        # monthly_paket_df = {'bulan': list(monthly_paket.keys()), 'omset_bulanan': list(monthly_paket.values())}
        # df_paket = pd.DataFrame(monthly_paket_df)
        # # End Penjualan Paket

        q4 = f"""
          {queryWith}
          SELECT m.bulan, IFNULL(SUM(dtp.harga_total), 0) AS omset_jual_produk, t.id_transaksi,
          dtp.is_addon, t.jenis_pembayaran
          FROM months m
          -- pakai left join lalu and ke main_transaksi
          LEFT JOIN main_transaksi t 
            ON DATE_FORMAT(t.created_at, '%Y-%m') = m.bulan 
            AND t.status IN ('paid', 'done') 
            AND t.is_cancel = 0
          -- pakai left join lalu and ke detail_trans
          LEFT JOIN detail_transaksi_produk dtp 
            ON t.id_transaksi = dtp.id_transaksi 
            AND dtp.status = 'paid'
          -- baru digroup by
          GROUP BY m.bulan, t.id_transaksi, dtp.is_addon, t.jenis_pembayaran
          ORDER BY m.bulan DESC
        """
        await cursor.execute(q4)
        items4 = await cursor.fetchall()

        # Buat list dlu supaya bs update data baru
        list_produk = []
        for data in items4:
          id_transaksi = data['id_transaksi']
          omset_jual_produk = int(data['omset_jual_produk']) or 0

          # ambil key maintrans yg sama dgn id_transaksi omset_jual_produk
          get_same_id = main_trans_dict.get(id_transaksi, {})
          disc = get_same_id.get('disc', 0)

          # Rumus Persen
          # if data['jenis_pembayaran'] == 1:
          #   # Case 1, Bayar Akhir, pasti dpt disc
          #   nominal_disc = omset_jual_produk * disc
          #   harga_stlh_disc = omset_jual_produk - nominal_disc
          # elif data['jenis_pembayaran'] == 0 and data['is_addon'] == 0:
          #   # Case 2: Payment upfront - only apply discount to non-addon items
          #   nominal_disc = omset_jual_produk * disc
          #   harga_stlh_disc = omset_jual_produk - nominal_disc
          # else:
          #   # Case 3: No discount applies (upfront payment for addon items)
          #   harga_stlh_disc = omset_jual_produk

          # Update key items3
          data['omset_jual_produk'] = omset_jual_produk
          list_produk.append(data)
        
        # Groupby Valuenya berdasarkan bulan
        monthly_produk = {}
        for item in list_produk:
          bulan = item['bulan']
          if bulan in monthly_produk:
            monthly_produk[bulan] += item['omset_jual_produk']
          else:
            monthly_produk[bulan] = item['omset_jual_produk']

        # Restruktur ulang dictionarynya lalu jadikan dataframe supaya membentuk kek sql
        monthly_produk_df = {'bulan': list(monthly_produk.keys()), 'omset_bulanan': list(monthly_produk.values())}
        df_produk = pd.DataFrame(monthly_produk_df) 

        # cek paket populer
        q5 = """
          SELECT 
            p.nama_paket_msg AS label,
            COUNT(*) AS jumlah_terjual
          FROM detail_transaksi_paket d
          JOIN paket_massage p ON d.id_paket = p.id_paket_msg
          WHERE d.status = 'paid'
            AND d.is_returned = 0
          GROUP BY d.id_paket
          ORDER BY jumlah_terjual DESC
          LIMIT 4;
        """
        await cursor.execute(q5)
        items5 = await cursor.fetchall()
        
        return {
          "for_line_chart" : items1,
          "monthly_sales": items2,
          "sum_paket": items3,
          "sum_produk": df_produk.to_dict('records'),
          "paket_terlaris": items5
        }

  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code) 
  
def excel_to_pdf(excel_path, pdf_path):
  excel = win32com.client.Dispatch("Excel.Application")
  excel.Visible = False #Buat Excel Hidden
  excel.DisplayAlerts = False #Lewati Alert

  try:
    print(f"Converting '{excel_path}' to PDF...")
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))
    # --- Add Page Setup to fit content ---
    ws = wb.ActiveSheet
    # --- Page Setup Configuration ---
    # 1. Set print area to only used cells
    used_range = ws.UsedRange
    ws.PageSetup.PrintArea = used_range.Address
    
    # 2. Fit to one page wide and tall
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1
    
    # 3. Prevent row/column splitting
    ws.PageSetup.FitToPagesTall = False  # Allow multiple pages if needed
    ws.PageSetup.Zoom = False  # Disable zoom to enforce FitToPages
    
    # 4. Set margins (optional, adjust as needed)
    ws.PageSetup.LeftMargin = 20
    ws.PageSetup.RightMargin = 20
    ws.PageSetup.TopMargin = 20
    ws.PageSetup.BottomMargin = 20
    
    # 5. Center on page
    ws.PageSetup.CenterHorizontally = True
    ws.PageSetup.CenterVertically = True
    
    # 6. Set paper size (A4)
    ws.PageSetup.PaperSize = 9  # 9 = xlPaperA4
    
    # 7. Set orientation (Auto: Excel will decide based on content)
    if used_range.Columns.Count > 10:  # If many columns, use landscape
        ws.PageSetup.Orientation = 2  # 2 = xlLandscape
    else:
        ws.PageSetup.Orientation = 1  # 1 = xlPortrait

    # --- Export to PDF ---
    wb.ActiveSheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    print("Conversion successful!")
        
  except com_error as e:
    print(f"Conversion failed: {e}")
  finally:
    if 'wb' in locals() and wb:
      wb.Close(SaveChanges=False)
    if excel:
      excel.Quit()
  
def formatStrDate(
  params: str
):
  tgl = params.split("-")
  formatted_tgl = tgl[2] + "-" + tgl[1] + "-" + tgl[0]
  return formatted_tgl

@app.get('/export_excel')
async def exportExcel(
  start_date: Optional[str] = Query(None),
  end_date: Optional[str] = Query(None)
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")

          kondisi = ""
          tgl = "" # buat munculin di excel. ini formatnya dd-mm-yyyy
          params= []
          if start_date and end_date:
            tgl = formatStrDate(start_date) + " s/d " + formatStrDate(end_date)
            # kalo utk kondisi formatny hrs sesuai dgn sql yaitu yyyy-mm-dd
            kondisi = "WHERE DATE(mt.created_at) BETWEEN %s AND %s AND mt.is_cancel = 0 AND mt.status in ('paid', 'done')"
            params.extend([start_date, end_date])
          elif start_date:
            tgl = formatStrDate(start_date)
            kondisi = "WHERE DATE(mt.created_at) = %s AND mt.is_cancel = 0 AND mt.status in ('paid', 'done')"
            params.append(start_date)            
          else:
            kondisi = "WHERE DATE(mt.created_at) = CURDATE()"
            # Ambil Date Now Skrg (sesuai server db)
            q_tgl = "SELECT NOW() AS tgl"
            await cursor.execute(q_tgl)

            item_tgl = await cursor.fetchone()
            fetched_tgl = str(item_tgl['tgl']).split(" ")
            tgl = formatStrDate(fetched_tgl[0])

          q1 = f"""
            SELECT ROW_NUMBER() OVER (ORDER BY mt.created_at) as no,mt.id_transaksi, mt.created_at AS tgl_beli, mt.jenis_transaksi, r.nama_ruangan AS kamar, 
            k_terapis.nama_karyawan AS terapis, CASE WHEN mt.jenis_pembayaran = 0 THEN 'pembayaran diawal' 
            ELSE 'pembayaran diakhir' END AS tipe_pembayaran,
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 0 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_paket dtp WHERE dtp.id_transaksi = mt. id_transaksi) AS INTEGER) + 
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 0 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_produk dtpr WHERE dtpr.id_transaksi = mt. id_transaksi) AS INTEGER) +
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 0 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_fnb dtfnb WHERE dtfnb.id_transaksi = mt. id_transaksi) AS INTEGER) AS Pembelian_Awal
            ,
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_paket dtp WHERE dtp.id_transaksi = mt. id_transaksi) AS INTEGER) + 
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_produk dtpr WHERE dtpr.id_transaksi = mt. id_transaksi) AS INTEGER) +
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_fnb dtfnb WHERE dtfnb.id_transaksi = mt. id_transaksi) AS INTEGER) AS Addon
            ,
            (mt.total_harga + 
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_paket dtp WHERE dtp.id_transaksi = mt. id_transaksi) AS INTEGER) + 
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_produk dtpr WHERE dtpr.id_transaksi = mt. id_transaksi) AS INTEGER) +
            CAST((SELECT COALESCE(SUM(CASE WHEN is_addon = 1 THEN harga_total ELSE 0 END), 0) FROM detail_transaksi_fnb dtfnb WHERE dtfnb.id_transaksi = mt. id_transaksi) AS INTEGER)) as total_harga, 
            mt.disc, mt.grand_total, mt.gtotal_stlh_pajak AS bayar
            FROM main_transaksi mt
            LEFT JOIN ruangan r ON mt.id_ruangan = r.id_ruangan
            -- JOIN tabel yang sama
            LEFT JOIN karyawan k ON mt.id_resepsionis = k.id_karyawan
            LEFT JOIN karyawan k_terapis ON mt.id_terapis = k_terapis.id_karyawan
            LEFT JOIN karyawan k_gro ON mt.id_gro = k_gro.id_karyawan
               {kondisi}
            GROUP BY 
            mt.id_transaksi, mt.created_at, mt.jenis_transaksi, 
            r.nama_ruangan, k_terapis.nama_karyawan, mt.jenis_pembayaran,
            mt.total_harga, mt.disc, mt.grand_total, mt.gtotal_stlh_pajak
          """
          
          await cursor.execute(q1, params)
          main_data = await cursor.fetchall()

          qdetail = """
            SELECT dtp.id_transaksi, pm.nama_paket_msg as nama_item, dtp.qty FROM detail_transaksi_paket dtp INNER JOIN paket_massage pm ON dtp.id_paket = pm.id_paket_msg
            UNION ALL
            SELECT dtp.id_transaksi, pe.nama_paket_extend as nama_item, dtp.qty FROM detail_transaksi_paket dtp INNER JOIN paket_extend pe ON dtp.id_paket = pe.id_paket_extend
            UNION ALL
            SELECT dtpr.id_transaksi, mp.nama_produk as nama_item, dtpr.qty FROM detail_transaksi_produk dtpr INNER JOIN menu_produk mp ON dtpr.id_produk = mp.id_produk
            UNION ALL
            SELECT dtpf.id_transaksi, mf.nama_fnb as nama_item, dtpf.qty FROM detail_transaksi_fnb dtpf INNER JOIN menu_fnb mf ON dtpf.id_fnb = mf.id_fnb
          """

          await cursor.execute(qdetail)
          detail_data = await cursor.fetchall()

          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Transaksi All"

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:M1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN PENJUALAN PERIODE {tgl}"
          ws.merge_cells('A2:M2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not main_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for {tgl}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          for col in main_data[0].keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          for trx in main_data:
            trx_id = trx['id_transaksi']
            row_as_str = []
            for value in trx.values():
              if isinstance(value, datetime):
                value = value.strftime('%d-%m-%Y\n%H:%M:%S')
              elif isinstance(value, int):
                value = value
              elif isinstance(value, float):
                if value == 0.0:
                  value = "-"
                else:
                  value = f"{int(value * 100)}%"

              elif value == "" or not value:
                value = "-"
              else:
                value = str(value)
              
              row_as_str.append(value)
            print(row_as_str)
            ws.append(row_as_str)

            ws.append([""] * (len(row_as_str)-12) + ["Detail Paket","Nama Item", "Qty"])

            current_row = ws.max_row
            start_col = len(row_as_str) - 12 + 1

            for col in range(start_col, start_col + 3):
              ws.cell(row=current_row, column=col).font = Font(bold=True)

            details = [d for d in detail_data if d['id_transaksi'] == trx_id]
            for d in details:
              ws.append([""] * (len(row_as_str)-11) + [d["nama_item"], str(d["qty"])])

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 12
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
              if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'


          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya

          summary_label = "TOTAL"
          total_bayar = sum(row.get('bayar',0)or 0 for row in main_data)
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", total_bayar])

          total_row_index = ws.max_row
          total_col_index = len(column_main)
          col_letter = get_column_letter(total_col_index)

          ws.merge_cells(f"{get_column_letter(total_col_index - 1)}{total_row_index}:{col_letter}{total_row_index}")


          # Style the summary row
          summary_row = ws[ws.max_row]  # Get the last row
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=total_row_index, column=total_col_index - 1)
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datapenjualan_platinum.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "datapenjualan_platinum.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="datapenjualan_platinum.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_bulanan')
async def export_excel_komisi_bulanan(
  month: Optional[str] = Query(None),
  year: Optional[str] = Query(None)
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([month, year, month, year, month, year])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi as CHAR)) <=3, pm.nominal_komisi/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s AND mt.status = 'done'

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi as CHAR)) <=3, mp.nominal_komisi/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi * dtpr.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s AND mp.nominal_komisi != 0 AND mt.status = 'done'
            
            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pe.nama_paket_extend as nama_paket, pe.harga_extend as harga_paket, dtp.qty, IF(LENGTH(CAST(pe.nominal_komisi as CHAR)) <=3, pe.nominal_komisi/100 * pe.harga_extend * dtp.qty , pe.nominal_komisi * dtp.qty ) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_extend as pe on dtp.id_paket = pe.id_paket_extend inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s AND mt.status = 'done'

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi Terapis & OB Bulanan "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI TERAPIS BULAN {month} TAHUN {year}"
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for month {month} and year {year}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Terapis :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '','', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 7 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)

            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
             
          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisiterapisbulanan.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_terapis_bulanan.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_terapis_bulanan.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_bulanan_gro')
async def export_excel_komisi_bulanan_gro(
  month: Optional[str] = Query(None),
  year: Optional[str] = Query(None)
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([month, year, month, year])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi_gro as CHAR)) <=3, pm.nominal_komisi_gro/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi_gro * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s AND pm.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtp.is_addon = 0

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi_gro as CHAR)) <=3, mp.nominal_komisi_gro/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi_gro * dtpR.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s AND mp.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtpr.is_addon = 0

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi GRO Bulanan "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI GRO BULAN {month} TAHUN {year}"
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for month {month} and year {year}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Gro :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '', '', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 6 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)
            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
              if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'

          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisigrobulanan.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_gro_bulanan.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_gro_bulanan.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_harian')
async def export_excel_komisi_harian(
  strdate: Optional[str] = Query(None),
  enddate: Optional[str] = Query(None)
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([strdate, enddate, strdate, enddate, strdate, enddate])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi as CHAR)) <=3, pm.nominal_komisi/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE DATE(mt.created_at) BETWEEN %s AND %s AND mt.status = 'done'

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi as CHAR)) <=3, mp.nominal_komisi/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi * dtpr.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE DATE(mt.created_at) BETWEEN %s AND %s AND mp.nominal_komisi != 0 AND mt.status = 'done'
            
            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pe.nama_paket_extend as nama_paket, pe.harga_extend as harga_paket, dtp.qty, IF(LENGTH(CAST(pe.nominal_komisi as CHAR)) <=3, pe.nominal_komisi/100 * pe.harga_extend * dtp.qty , pe.nominal_komisi * dtp.qty ) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_extend as pe on dtp.id_paket = pe.id_paket_extend inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE DATE(mt.created_at) BETWEEN %s AND %s AND mt.status = 'done'

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi Terapis & OB Bulanan "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI TERAPIS Tanggal {strdate} - Tanggal {enddate}"
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for Tanggal {strdate} - Tanggal {enddate}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Terapis :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '','', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 7 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)
              
            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
             
          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisiterapisharian.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_terapis_harian.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_terapis_harian.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_harian_gro')
async def export_excel_komisi_harian_gro(
  strdate: Optional[str] = Query(None),
  enddate: Optional[str] = Query(None)
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([strdate, enddate, strdate, enddate])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi_gro as CHAR)) <=3, pm.nominal_komisi_gro/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi_gro * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE DATE(mt.created_at) BETWEEN %s AND %s AND pm.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtp.is_addon = 0

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi_gro as CHAR)) <=3, mp.nominal_komisi_gro/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi_gro * dtpR.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE DATE(mt.created_at) BETWEEN %s AND %s AND mp.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtpr.is_addon = 0

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi GRO Harian "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI GRO Tanggal {strdate} - Tanggal {enddate}"
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for Tanggal {strdate} and Tanggal {enddate}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Gro :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '','', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 6 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)
            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
              if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'

          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisigroharian.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_gro_harian.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_gro_harian.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_tahunan')
async def export_excel_komisi_tahunan(
  year: Optional[str] = Query(None),
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([year,year,year])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi as CHAR)) <=3, pm.nominal_komisi/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE YEAR(mt.created_at) = %s AND mt.status = 'done'

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi as CHAR)) <=3, mp.nominal_komisi/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi * dtpr.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE YEAR(mt.created_at) = %s AND mp.nominal_komisi != 0 AND mt.status = 'done'
            
            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pe.nama_paket_extend as nama_paket, pe.harga_extend as harga_paket, dtp.qty, IF(LENGTH(CAST(pe.nominal_komisi as CHAR)) <=3, pe.nominal_komisi/100 * pe.harga_extend * dtp.qty , pe.nominal_komisi * dtp.qty ) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_extend as pe on dtp.id_paket = pe.id_paket_extend inner join karyawan as k on mt.id_terapis = k.id_karyawan WHERE YEAR(mt.created_at) = %s AND mt.status = 'done'

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi Terapis & OB Tahunan "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI TERAPIS Tahun {year}"
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for Tahun {year}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Terapis :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '','', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 7 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)
              
            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
             
          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisiterapistahunan.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_terapis_tahunan.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_terapis_tahunan.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

@app.get('/export_excel_komisi_tahunan_gro')
async def export_excel_komisi_tahunan_gro(
  year: Optional[str] = Query(None),
):
  try :
    pool = await get_db()

    async with pool.acquire() as conn:
      async with conn.cursor(aiomysql.DictCursor) as cursor:
        try:
          await cursor.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED;")
          kondisi = ""
          params= []
          # kondisi = "MONTH(mt.created_at) = %s and YEAR(mt.created_at) = %s"
          params.extend([year, year])

          print(params)
          q1 = """

            SELECT mt.id_transaksi, mt.created_at, dtp.id_paket as id_paket, pm.nama_paket_msg as nama_paket, pm.harga_paket_msg as harga_paket, dtp.qty, IF(LENGTH(CAST(pm.nominal_komisi_gro as CHAR)) <=3, pm.nominal_komisi_gro/100 * pm.harga_paket_msg * dtp.qty, pm.nominal_komisi_gro * dtp.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_paket as dtp on mt.id_transaksi = dtp.id_transaksi 
            inner join paket_massage as pm on dtp.id_paket = pm.id_paket_msg inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE YEAR(mt.created_at) = %s AND pm.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtp.is_addon = 0

            UNION ALL

            SELECT mt.id_transaksi, mt.created_at, dtpr.id_produk as id_paket, mp.nama_produk as nama_paket, mp.harga_produk as harga_paket, dtpr.qty, IF(LENGTH(CAST(mp.nominal_komisi_gro as CHAR)) <=3, mp.nominal_komisi_gro/100 * mp.harga_produk * dtpr.qty, mp.nominal_komisi_gro * dtpR.qty) AS komisi , k.nama_karyawan 
            FROM main_transaksi as mt inner join detail_transaksi_produk as dtpr on mt.id_transaksi = dtpr.id_transaksi 
            inner join menu_produk as mp on dtpr.id_produk = mp.id_produk inner join karyawan as k on mt.id_gro = k.id_karyawan WHERE YEAR(mt.created_at) = %s AND mp.nominal_komisi_gro != 0 AND mt.status = 'done' AND dtpr.is_addon = 0

            ORDER BY nama_karyawan, id_transaksi ASC

          """

          print(q1)
          await cursor.execute(q1, params)
          main_data_komisi = await cursor.fetchall()

          print('data :',main_data_komisi)
          # Step 2, buat Workbook Excel. wb = WorkBook, ws = WorkSheet
          wb = Workbook()
          ws = wb.active
          ws.title = "Komisi GRO Tahunan "

          grouped_data = defaultdict(list)
          for row in main_data_komisi:
            grouped_data[row['nama_karyawan']].append(row)

          # Step 3, Tambahkan Header kaya Judul lalu di Merge
          corporate_name = "PLATINUM"
          ws.merge_cells('A1:H1') # Merge A1 smpe I1
          corp_cell = ws['A1']
          corp_cell.value = corporate_name
          ws.row_dimensions[2].height = 30
          ws.row_dimensions[1].height = 30

          # Step 4, Center Text Corporate Name
          corp_cell.alignment = Alignment(horizontal='center', vertical='center')
          header_font = Font(bold=True, size=18)
          for cell in ws[1]:# index row excel start dr 1. ini ceritany mw tebalin PLATINUM
            cell.font = header_font

          # Step 5, Tambah Keterangan dibawah nama korporat yg diheader A1
          corporate_ket = f"LAPORAN KOMISI GRO Tahun {year} "
          ws.merge_cells('A2:H2')
          ket_cell = ws['A2']
          ket_cell.value = corporate_ket

          # Step 6. Center Tulisan LapPenjualan
          ket_cell.alignment = Alignment(horizontal='center', vertical='center')
          header2_font = Font(bold=True, size=15)
          for cell in ws[2]:
            cell.font = header2_font

          # Buat Row Kosong
          ws.append([])

          # Check if data is empty
          if not grouped_data:
            print("Transaksi Kosong")
            return JSONResponse(
              {"message": f"No transaction data found for Tahun {year}"},
              status_code=500
            )

          # Step 7 : Tambah Header. konversi ke string dlu
          column_main = []
          first_row = list(grouped_data.values())[0][0]
          #jadiin first_row untuk ambil keysnya
          for col in first_row.keys():
            replaced_str = str(col).capitalize().replace("_", " ")
            column_main.append(replaced_str)

          ws.append(column_main)

          # Step 7.5 Desain Header. ws itu adalah worksheet row ke 4
          header_row = ws[4] # Diambil dari row ke 4 yg udh di append pada step 7.5
          for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )
          
          for karyawan, rows in grouped_data.items():
            ws.append(['Gro :', karyawan])

            total_komisi = 0
            for r in rows :
              formatted_komisi = "{:,.0f}".format(r['komisi']).replace(",", ".")
              formatted_harga_paket = "{:,.0f}".format(r['harga_paket']).replace(",", ".")
              formatted_time = r['created_at'].strftime('%d-%m-%Y %H:%M:%S')

              ws.append([
                r['id_transaksi'],
                formatted_time,
                r['id_paket'],
                r['nama_paket'],
                formatted_harga_paket,
                r['qty'],
                formatted_komisi,
                r['nama_karyawan']
              ])

              last_row = ws.max_row
              harga_cell =  ws.cell(row=last_row, column=5)
              harga_cell.alignment = Alignment(horizontal='right')

              komisi_cell =  ws.cell(row=last_row, column=7)
              komisi_cell.alignment = Alignment(horizontal='right')

              qty_cell = ws.cell(row=last_row, column=6)
              qty_cell.alignment = Alignment(horizontal='center',vertical='center')
              
              total_komisi += r['komisi']
              formatted_total_komisi = "{:,.0f}".format(total_komisi).replace(",", ".")
            
            ws.append(['TOTAL', '', '', '', '','', formatted_total_komisi, ''])
            total_row_idx = ws.max_row #This points to the TOTAL row just added


            for cell in ws[total_row_idx]:  # Second-to-last row is the "TOTAL" row
              cell.font = Font(bold=True)
              cell.alignment = Alignment(horizontal = 'right')  if cell.column == 6 else Alignment(horizontal = 'center')
              ws.merge_cells(start_column=1,start_row= total_row_idx,end_row=total_row_idx, end_column=5)
            ws.append([])

          # Step 8 : Tambah data ke ws. jadikan list dlu, br ambil values
          # for row in grouped_data.values():
          #   for row in rows:
          #     row_as_str = []
          #     for value in row:
          #       if isinstance(value, datetime):
          #         value = value.strftime('%d-%m-%Y\n%H:%M:%S')
          #       elif isinstance(value, int):
          #         value = value
          #       elif isinstance(value, float):
          #         value = str(value).replace("0.", "")
          #         if value == "0":
          #           value = "-"
          #         else:
          #           value += "0%"

          #       elif value == "" or not value:
          #         value = "-"
          #       else:
          #         value = str(value)
                
          #       row_as_str.append(value)
          #     print(row_as_str)
          #     ws.append(row_as_str)

            # Step 8.5: Tambahkan border ke setiap row
            thin_border = Border(
              bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row):  # Mulai dari row 5 (data pertama setelah header)
              for cell in row:
                cell.border = thin_border

          # Step 9 : Auto Adjust Column width berdasarkan konten
          for idx, col in enumerate(ws.columns, 1):  # Mulai dari 1 (kolom A)
            column_letter = get_column_letter(idx)
            max_length = 0
            
            for cell in col:
              # Lewati looping yg mergedcell
              if isinstance(cell, MergedCell):
                continue
              
              try:
                # Handle khusus untuk teks yang ada newline
                cell_value = str(cell.value)

                # Penanganan khusus untuk kolom id_transaksi (kolom pertama)
                if idx == 1:  # Kolom A (id_transaksi)
                  # Tetapkan panjang maksimum 8 karakter untuk id_transaksi
                  max_length = 8
                  break  # Keluar dari loop karena kita sudah tetapkan manual

                elif "\n" in cell_value:
                  # ambil line terpanjang
                  line_lengths = [len(line) for line in cell_value.split("\n")]
                  cell_length = max(line_lengths)
                else:
                  # buat lebar cell berdasarkan panjang data
                  cell_length = len(cell_value)

                if cell_length > max_length:
                  max_length = len(str(cell.value))
              except:
                pass
            
            if idx == 1:
              # Khusus id_transaksi ttpin 15 karakter
              adjusted_width = 15
            elif idx == 2:
              # Tglbeli, 12 karakter aja
              adjusted_width = 20
            else:
              # Kasih Buffer 20% dan minimum width 5
              adjusted_width = max((max_length + 2) * 1.2, 5)

            ws.column_dimensions[column_letter].width = adjusted_width

            for cell in col:
              # aktifkan wraptext untuk kolom yang ada newline
              if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True)
              # Format yg Bersifat Int
              if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'

          # After adding all your data rows (after the for row in main_data loop)
          # Add an empty row for separation
          ws.append([])

          # Add the summary row. adjust yang len(column_main) - 3. klo mw perkecil, besarin  valuenya
          summary_label = "TOTAL SELURUH KOMISI"
          total_komisi = sum(float(row['komisi']) for rows in grouped_data.values() for row in rows)
          formattedkomisi = "{:,.0f}".format(total_komisi).replace(",", ".")
          
          ws.append([summary_label] + [""] * (len(column_main) - 4) + ["", formattedkomisi])

          # Style the summary row
          summary_row_idx = ws.max_row  # Get the last row
          ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=2)

          summary_row = ws[summary_row_idx]
          for cell in summary_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical= 'center')
            cell.border = Border(
              left=Side(style="thin"),
              right=Side(style="thin"),
              top=Side(style="thin"),
              bottom=Side(style="thin"),
            )

          # Make the total value right-aligned and formatted with thousands separator
          total_cell = ws.cell(row=ws.max_row, column=len(column_main))
          total_cell.alignment = Alignment(horizontal="right")
          total_cell.number_format = '#,##0'

          # Step 10 : Simpan Workbook Ke File
          file_path = "datakomisigrotahunan.xlsx"
          # Temporarily Make File Writable Before Overwriting
          if os.path.exists(file_path):
            os.chmod(file_path, 0o644)  # Grant write permissions (rw-r--r--)

          wb.save(file_path)
          os.chmod(file_path, 0o444)  # Set back to read-only

          # Panggil Fungsi excel_to_pdf yg manual aku buat
          pdf_output = "data_komisi_gro_tahunan.pdf"
          excel_to_pdf(file_path, pdf_output)

          await asyncio.sleep(1.0)

          # # Step 11. Return Excelny sbg FileResponse
          return FileResponse(
            os.path.abspath(pdf_output),
            media_type='application/pdf',
            filename="data_komisi_gro_tahunan.pdf"
          )

        except aiomysqlerror as e:
          return JSONResponse({"Error": f"Sql Error {str(e)}"}, status_code=500)
        except HTTPException as e:
          return JSONResponse({"Error": f"HTTP Error {str(e.detail)}"}, status_code=e.status_code)
  except HTTPException as e:
   return JSONResponse({"Error": str(e)}, status_code=e.status_code)

