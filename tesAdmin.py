import json

from openpyxl import Workbook
from fastapi import FastAPI, File, Form, UploadFile, APIRouter, Request, HTTPException, Security, Query, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi_jwt import (
  JwtAccessBearerCookie,
  JwtAuthorizationCredentials,
  JwtRefreshBearer
)
import pandas as pd
from jwt_auth import access_security
from koneksi import conn
import os
import uuid
from datetime import datetime
from typing import List, Optional, Union
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.fnConvertStr import serialize_data
from utils.fnWebScrapping import scraper

app = APIRouter()

#get date now - 1 bln
def getDateOneMonthAgo():
  now = datetime.now()

  one_month_ago = now - relativedelta(months=1)

  return str(one_month_ago)

# @app.get('/data-scraping')
# def scrappedData():
#   data = scraper("https://www.traveloka.com/id-id/explore/activities/wisata-pontianak-yang-lagi-hits-ta/321888")

#   return data


#Bagian Dashboard
@app.get('/dashboard')
async def getDashboard() :
  try:
    cursor = conn.cursor()

    q1 = """
      SELECT COUNT(email) AS totalRegistered FROM users WHERE roles != %s AND 
      created_at >= NOW() - INTERVAL '1 month'
    """
    cursor.execute(q1, ('ADMIN', ))
    column_name = []
    for kol in cursor.description:
      column_name.append(kol[0])
    items1 = cursor.fetchall()

    q2 = """
      SELECT SUM(total_harga) AS totalPendapatan FROM transaksi WHERE status_trans = %s
      AND tgl_trans >= NOW() - INTERVAL '1 month'
    """
    cursor.execute(q2, ('COMPLETED', ))
    column_name2 = [kol[0] for kol in cursor.description]
    items2 = cursor.fetchall()

    q3 = """
      SELECT COUNT(id_trans) AS totalTiket FROM transaksi WHERE status_trans = %s
      AND tgl_trans >= NOW() - INTERVAL '1 month'
    """
    cursor.execute(q3, ('COMPLETED', ))
    column_name3 = [kol[0] for kol in cursor.description]
    items3 = cursor.fetchall()

    q4 = """
      SELECT t.*, u.username, u.profile_picture FROM transaksi t
      INNER JOIN users u
      ON t.email_cust = u.email
      ORDER BY t.tgl_trans DESC LIMIT 5
    """ 
    cursor.execute(q4)
    column_name4 = [kol[0] for kol in cursor.description]
    items4 = cursor.fetchall()


    # Ambil 6 Bln Terakhir, mainkan di generate seriesnya
    q5 = """
      SELECT 
        DATE_TRUNC('month', NOW()) - INTERVAL '1 month' * gs.month AS month_start,
        COUNT(t.id_trans) AS totalTiket
      FROM 
        generate_series(0, 5) AS gs(month)
      LEFT JOIN 
        transaksi t ON DATE_TRUNC('month', t.tgl_trans) = DATE_TRUNC('month', NOW()) - INTERVAL '1 month' * gs.month
        AND t.status_trans = 'COMPLETED'
      GROUP BY 
        month_start
      ORDER BY 
        month_start DESC;
    """
    cursor.execute(q5)
    column_name5 = [kol[0] for kol in cursor.description]
    items5 = cursor.fetchall()


    data1 = pd.DataFrame(items1, columns=column_name)
    data2 = pd.DataFrame(items2, columns=column_name2)
    data3 = pd.DataFrame(items3, columns=column_name3)
    data4 = pd.DataFrame(items4, columns=column_name4)
    data5 = pd.DataFrame(items5, columns=column_name5)

    # Pecah arraynya dulu, abis itu ambil key object totalpendapatan yg dihasilkan oleh sql
    convColumnData2 = data2.to_dict('records')[0]['totalpendapatan']

    # Pecah yg data4 skrg.
    convColumnData4 = []
    
    for i in data4.to_dict('records'):
      i['total_harga'] = f"{i['total_harga']:,.2f}"
      convColumnData4.append(i)

    # :, sebagai separator bilangan ribuan
    # :.2f buat nambah 2 angka di blkg comma 


    return {
      "dataRegistered": data1.to_dict('records')[0],
      "dataPendapatan": f"{convColumnData2:,.2f}", #convertstring ala python
      "dataTiket": data3.to_dict('records')[0],
      "dataTrans": convColumnData4,
      "previousDate": getDateOneMonthAgo().split(" "),
      "sixMonthDate": data5.to_dict('records')
    }


  except Exception as e:
    print(e)
  finally:
    cursor.close()


#Bagian Rute
@app.get('/rute')
async def getRute(
  id_rute: Optional[str] = Query(None),
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()
    if id_rute is None:
      q1 = "SELECT * FROM rute"
      cursor.execute(q1)
    else:
      q1 = "SELECT * FROM rute WHERE id = %s"
      cursor.execute(q1, (id_rute, ))

    items = cursor.fetchall()

    column_name = []
    for kol in cursor.description:
      column_name.append(kol[0])

    df = pd.DataFrame(items, columns=column_name)
    return df.to_dict('records')

  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  
  finally:
    cursor.close()

@app.post('/insertRute')
async def insertRute(
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()
    data = await request.json()

    id = data['id']
    kotaAwal = data['kota_awal']
    kotaAkhir = data['kota_akhir']
    waktuBrkt = data['waktu_berangkat']
    waktuSampai = data['waktu_sampai']
    lamaTempuh = data['lama_tempuh_jam']

    q1 = """
      INSERT INTO rute VALUES(%s, %s, %s, %s, %s, %s)
    """
    cursor.execute(q1, (id, kotaAwal, kotaAkhir, lamaTempuh, waktuBrkt, waktuSampai))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  
  except HTTPException as e:
    return JSONResponse(content={"Error": "Dat tdk simpan"}, status_code=e.status_code)
  finally:
    cursor.close()

@app.put('/editRoute/{id}')
async def editRute(
  id: str,
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()
    data = await request.json()

    kotaAwal = data['kota_awal']
    kotaAkhir = data['kota_akhir']
    waktuBrkt = data['waktu_berangkat']
    waktuSampai = data['waktu_sampai']
    lamaTempuh = data['lama_tempuh_jam']

    q1 = """
      UPDATE rute SET kota_awal = %s, kota_akhir = %s, waktu_berangkat = %s, waktu_sampai = %s, lama_tempuh_jam = %s
      WHERE id = %s
    """
    cursor.execute(q1, (kotaAwal, kotaAkhir, waktuBrkt, waktuSampai, lamaTempuh, id))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  
  except HTTPException as e:
    return JSONResponse(content={"Error": "Dat tdk simpan"}, status_code=e.status_code)
  finally:
    cursor.close()


@app.delete('/rute/{id}')
async def deleteRute(
  id: str,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()

    q1 = """
      DELETE FROM rute WHERE id = %s
    """
    cursor.execute(q1, (id, ))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Kehapus"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()
#End Bagian Rute

#yang getnya ada di ruteTransaksi.py
#Start bagian insertBis
@app.post('/insertBis')
async def insertBis(
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.form()

    id = data.get('id_bis')
    namaBis = data.get('nama_bis')
    idKelasBis = data.get('id_kelas_bis')
    jasaTravel = data.get('jasa_travel')
    idRute = data.get('id_rute')
    logojasatravel = data.get('logojasatravel')
    listGbrBis: List[UploadFile] = data.getlist('gbrbis')

    # Ini buat yg Logo
    filenameLogo = f"{uuid.uuid4()}.jpg"
    file_location_logo = os.path.join("images/logoBis/", filenameLogo)

    #read & save file logo 
    contentLogo = await logojasatravel.read()
    with open(file_location_logo, "wb") as f:
      f.write(contentLogo)
    # End Buat Yg Logo

    # Ini Buat yg gbrBis
    uploadedListBis = []
    for file in listGbrBis:
      file.filename = f"{uuid.uuid4()}.jpg"
      content = await file.read()

      # Simpan Filenya
      with open(f"images/gbrbis/{file.filename}", "wb") as f:
        f.write(content)

      uploadedListBis.append(file.filename)

    # convert list ke json formatted string
    gambar_json = json.dumps(uploadedListBis)

    # End Buat GbrBis

    q1 = """
      INSERT INTO bis VALUES(%s, %s, %s, %s, %s, %s, %s)
    """
    cursor.execute(q1, (id, namaBis, idKelasBis, jasaTravel, idRute, filenameLogo, gambar_json))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()

@app.put('/updateBis/{id}')
async def updateBis(
  id:str,
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.form()

    namaBis = data.get('nama_bis')
    idKelasBis = data.get('id_kelas_bis')
    jasaTravel = data.get('jasa_travel')
    idRute = data.get('id_rute')
    logojasatravel = data.get('logojasatravel')
    listGbrBis: List[UploadFile] = data.getlist('gbrbis')

    if logojasatravel:
      # Ini buat yg Logo
      filenameLogo = f"{uuid.uuid4()}.jpg"
      file_location_logo = os.path.join("images/logoBis/", filenameLogo)

      #read & save file logo 
      contentLogo = await logojasatravel.read()
      with open(file_location_logo, "wb") as f:
        f.write(contentLogo)
      # End Buat Yg Logo

    if listGbrBis:
      # Ini Buat yg gbrBis
      uploadedListBis = []
      for file in listGbrBis:
        file.filename = f"{uuid.uuid4()}.jpg"
        content = await file.read()

        # Simpan Filenya
        with open(f"images/gbrbis/{file.filename}", "wb") as f:
          f.write(content)

        uploadedListBis.append(file.filename)

      # convert list ke json formatted string
      gambar_json = json.dumps(uploadedListBis)
      # End Buat GbrBis

    
    if logojasatravel and listGbrBis:
      q1 = """
        UPDATE bis SET nama_bis = %s, id_kelas_bis = %s, jasa_travel = %s, id_rute = %s,
        logojasatravel = %s, gbrbis = %s WHERE id_bis = %s
      """
      cursor.execute(q1, (namaBis, idKelasBis, jasaTravel, idRute, filenameLogo, gambar_json, id))

    elif listGbrBis:
      q1 = """
        UPDATE bis SET nama_bis = %s, id_kelas_bis = %s, jasa_travel = %s, id_rute = %s,
        gbrbis = %s WHERE id_bis = %s
      """
      cursor.execute(q1, (namaBis, idKelasBis, jasaTravel, idRute, gambar_json, id))

    elif logojasatravel:
      q1 = """
        UPDATE bis SET nama_bis = %s, id_kelas_bis = %s, jasa_travel = %s, id_rute = %s,
        logojasatravel = %s WHERE id_bis = %s
      """
      cursor.execute(q1, (namaBis, idKelasBis, jasaTravel, idRute, filenameLogo, id))

    else:
      q1 = """
        UPDATE bis SET nama_bis = %s, id_kelas_bis = %s, jasa_travel = %s, id_rute = %s
        WHERE id_bis = %s
      """
      cursor.execute(q1, (namaBis, idKelasBis, jasaTravel, idRute, id))

    conn.commit()
    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()

@app.delete('/listbis/{id}')
async def deleteBis(
  id: str,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()

    q1 = """
      DELETE FROM bis WHERE id_bis = %s
    """
    cursor.execute(q1, (id, ))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Kehapus"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()
#End Bagian insertBis

# Start Stok Tiket
@app.get('/stoktiket')
async def getStokTiket(
  id_bis: Optional[str] = Query(None)
):
  try:
    if id_bis is None:
      cursor = conn.cursor()
      q1 = """
        SELECT s.*, b.nama_bis FROM stok_tiket s
        INNER JOIN bis b ON s.id_bis = b.id_bis
      """
      cursor.execute(q1)
    else:
      cursor = conn.cursor()
      q1 = """
        SELECT s.*, b.nama_bis FROM stok_tiket s
        INNER JOIN bis b ON s.id_bis = b.id_bis
        WHERE s.id_bis = %s
      """
      cursor.execute(q1, (id_bis, ))

    items = cursor.fetchall()

    column_name = []
    for kol in cursor.description:
      column_name.append(kol[0])

    df = pd.DataFrame(items, columns=column_name)
    return df.to_dict('records')

  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  
  finally:
    cursor.close()

@app.post('/stoktiket')
async def insertStokTiket(
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.json()

    id = data['id_bis']
    dt1 = data['total_tiket']
    dt2 = data['tiket_tersedia']

    q1 = """
      INSERT INTO stok_tiket VALUES(%s, %s, %s)
    """
    cursor.execute(q1, (id, dt1, dt2))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()

@app.put('/stoktiket/{id}')
async def updateStokTiket(
  id:str,
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.json()

    dt1 = data['total_tiket']
    dt2 = data['tiket_tersedia']

    q1 = """
      UPDATE stok_tiket SET total_tiket = %s, tiket_tersedia = %s
      WHERE id_bis = %s
    """
    cursor.execute(q1, (dt1, dt2, id))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()
# End Bagian Stok Tiket


#Start Bagian Kelas Bis
@app.get('/kelasbis')
async def getKelasBis(
  id_kelas: Optional[str] = Query(None)
):
  try:
    if id_kelas is None:
      cursor = conn.cursor()
      q1 = "SELECT * FROM kelas_bis"
      cursor.execute(q1)
    else:
      cursor = conn.cursor()
      q1 = "SELECT * FROM kelas_bis WHERE id_kelas = %s"
      cursor.execute(q1, (id_kelas, ))

    items = cursor.fetchall()

    column_name = []
    for kol in cursor.description:
      column_name.append(kol[0])

    df = pd.DataFrame(items, columns=column_name)
    return df.to_dict('records')

  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  
  finally:
    cursor.close()

@app.put('/kelasbis/{id}')
async def updateKelasBis(
  id:str,
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.json()

    dt1 = data['nama_kelas']
    dt2 = data['harga']

    q1 = """
      UPDATE kelas_bis SET nama_kelas = %s, harga = %s
      WHERE id_kelas = %s
    """
    cursor.execute(q1, (dt1, dt2, id))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()


@app.post('/kelasbis')
async def insertKelasBis(
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  try:
    cursor = conn.cursor()

    data = await request.json()

    id = data['id_kelas']
    dt1 = data['nama_kelas']
    dt2 = data['harga']

    q1 = """
      INSERT INTO kelas_bis VALUES(%s, %s, %s)
    """
    cursor.execute(q1, (id, dt1, dt2))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Tersimpan"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()

@app.delete('/kelasbis/{id}')
async def deleteKelasBis(
  id: str,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()

    q1 = """
      DELETE FROM kelas_bis WHERE id_kelas = %s
    """
    cursor.execute(q1, (id, ))
    conn.commit()

    return JSONResponse(content={"Success": "Dat Kehapus"}, status_code=200)
  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()
#End Bagian KelasBis

# Start Transaksi
@app.get('/dataTrans')
async def getAllTrans(
  id: Optional[str] = Query(None), # Untuk ambil queryString
  user: JwtAuthorizationCredentials = Security(access_security)
):
  if user['roles'] == "ADMIN" or user['roles'] == "KASIR":
    try:
      cursor = conn.cursor()

      if id is not None:
        q1 = """
          SELECT t.*, u.email, u.username, dt.id_bis, dt.tgl_pergi, dt.jlh_penumpang, dt.tgl_balik, 
          b.id_rute, b.nama_bis, b.id_kelas_bis, pw.nama_paket, r.kota_awal, r.kota_akhir FROM transaksi t 
          INNER JOIN "detailTransaksi" dt ON t.id_trans = dt.id_trans
          INNER JOIN bis b ON dt.id_bis = b.id_bis
          INNER JOIN users u ON t.email_cust = u.email
          LEFT JOIN rute r ON b.id_rute = r.id
          LEFT JOIN paketwisata pw ON t.id_paket = pw.id_paket
          WHERE t.id_trans = %s
        """
        cursor.execute(q1, (id, )) #kalo tuple sifatnya gini. klo data single kasih 1 koma ksg
      else:
        q1 = """
          SELECT t.*, dt.tgl_pergi, dt.tgl_balik, b.id_rute, pw.nama_paket FROM transaksi t 
          INNER JOIN "detailTransaksi" dt ON t.id_trans = dt.id_trans
          INNER JOIN bis b ON dt.id_bis = b.id_bis
          LEFT JOIN paketwisata pw ON t.id_paket = pw.id_paket
        """
        cursor.execute(q1)

      column_name = []
      for kol in cursor.description:
        column_name.append(kol[0])
      
      items = cursor.fetchall()

      df = pd.DataFrame(items, columns=column_name)

      data = df.to_dict('records')
      return data

    except HTTPException as e:
      return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
    
    finally:
      cursor.close()
  else:
    return JSONResponse(content={"Error": "Unauhorized"}, status_code=401)


@app.get('/filterDataTrans/{mode}')
async def getFilterTrans(
  mode: str,
  user: JwtAuthorizationCredentials = Security(access_security)
):
  if user['roles'] == "ADMIN" or user['roles'] == "KASIR":
    try:
      cursor = conn.cursor()

      q1 = f"""
          SELECT t.*, dt.tgl_pergi, dt.tgl_balik, b.id_rute, pw.nama_paket FROM transaksi t 
          INNER JOIN "detailTransaksi" dt ON t.id_trans = dt.id_trans
          INNER JOIN bis b ON dt.id_bis = b.id_bis
          LEFT JOIN paketwisata pw ON t.id_paket = pw.id_paket
          {"WHERE t.status_trans = 'COMPLETED'" if mode == 'COMPLETED' else ''}
          {"WHERE t.status_trans = 'PENDING'" if mode == 'PENDING' else ''}
          {"WHERE t.status_trans = 'CANCELLED'" if mode == 'CANCELLED' else ''}
          {"ORDER BY t.tgl_trans DESC" if mode == 'TBARU' else ''}
          {"ORDER BY t.tgl_trans ASC" if mode == 'TLAMA' else ''}
      """
      cursor.execute(q1)

      column_name = []
      for kol in cursor.description:
        column_name.append(kol[0])
      
      items = cursor.fetchall()

      df = pd.DataFrame(items, columns=column_name)

      data = df.to_dict('records')
      return data

    except HTTPException as e:
      return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
    
    finally:
      cursor.close()
  else:
    return JSONResponse(content={"Error": "Unauhorized"}, status_code=401)


admin_connection = []

@app.websocket("/ws-transaksi")
async def wsTransaksi(
  websocket: WebSocket
):
  await websocket.accept()
  admin_connection.append(websocket)
  try:
    await websocket.send_text(json.dumps({"message": "Hello from ws Transaksi"}))

    while True:
      print("Sudah Konek Ws Transaksi")
      # Biarkan Koneksi Tetap Nyala
      await websocket.receive_text()
  except WebSocketDisconnect:
    admin_connection.remove(websocket)

@app.put('/dataTrans/{id}')
async def updateTrans(
  id: str,
  request: Request,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  cursor = conn.cursor()
  try:
    cursor.execute("START TRANSACTION")
    data = await request.json()
    # print(data)

    if data['status_trans'] == "CANCELLED":
      q1 = "UPDATE transaksi SET status_trans = %s, id_staff = %s, alasan_tolak = %s WHERE id_trans = %s"
      cursor.execute(q1, (data['status_trans'], data['id_staff'], data['alasan_tolak'], id))

      # Balikin Lagi Stoknya
      qSelect = """
        SELECT jlh_penumpang, id_bis FROM "detailTransaksi" WHERE id_trans = %s
      """
      cursor.execute(qSelect, (id, ))
      result = cursor.fetchone()

      q2 = "UPDATE stok_tiket SET tiket_tersedia = tiket_tersedia + %s WHERE id_bis = %s"
      cursor.execute(q2, (result[0], result[1]))
    else:
      if data['metode_byr'] == "transfer":
        # Select total_harga untuk user yg bayar make transfer. samain total harga dgn user bayar
        qSelect = "SELECT total_harga FROM transaksi WHERE id_trans = %s"
        cursor.execute(qSelect, (id, ))
        result = cursor.fetchone()

        q1 = "UPDATE transaksi SET status_trans = %s, id_staff = %s, generated_ticket = %s, user_bayar = %s, kembalian = %s WHERE id_trans = %s"
        cursor.execute(q1, (data['status_trans'], data['id_staff'], data['generated_ticket'], result[0], 0, id))
      else:
        q1 = "UPDATE transaksi SET status_trans = %s, id_staff = %s, generated_ticket = %s, user_bayar = %s, kembalian = %s WHERE id_trans = %s"
        cursor.execute(q1, (data['status_trans'], data['id_staff'], data['generated_ticket'], data['user_bayar'], data['kembalian'], id))

    conn.commit()

    qTransUser = "SELECT * FROM transaksi WHERE id_trans = %s"
    cursor.execute(qTransUser, (id, ))
    items = cursor.fetchone() 

    data['id_trans'] = id
    data['email_cust'] = items[3] #Email User di index ke 3

    # Kirim Data ke Websocket
    for con in admin_connection:
      await con.send_text(json.dumps(data))

    return JSONResponse(content=data, status_code=200)

  except HTTPException as e:
    cursor.execute("ROLLBACK")
    return JSONResponse(content={"ErrorWoi": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()


@app.delete('/dataTrans/{id}')
async def updateTrans(
  id: str,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  cursor = conn.cursor()

  try:
    cursor.execute("START TRANSACTION")
    # print(data)

    q1 = "DELETE FROM transaksi WHERE id_trans = %s"
    cursor.execute(q1, (id, ))

    q2 = """
      DELETE FROM "detailTransaksi" WHERE id_trans = %s
    """
    cursor.execute(q2, (id, ))

    conn.commit()
    return JSONResponse(content={"Success": "Lala"}, status_code=200)

  except HTTPException as e:
    cursor.execute("ROLLBACK")
    return JSONResponse(content={"ErrorWoi": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()


#End Transaksi
  
  
# Paket. yg cekpaket ada di routeTransaction.py
def idPaket():
  try:
    cursor = conn.cursor()
    q1 = "SELECT id_paket FROM paketwisata ORDER BY id_paket DESC LIMIT 1"
    cursor.execute(q1)

    id_awal = cursor.fetchone() #ini haslnya bentuk array. di pecahkan dlu

    if len(id_awal) == 0:
      keyword = "P0001"
      return keyword
    else:
      idAwalToStr = id_awal[0]
      substrIdAwal = idAwalToStr[1:5] #ambil str di index 1, sampai ke "karakter ke 5"
      toIntId = int(substrIdAwal) + 1
      keyword = "P" + str(toIntId).zfill(4) #semacam padleft, dia nambahin zero didepan.
      return keyword
    
  except Exception as e:
    print(e)
  finally:
    cursor.close()


def fnIdBenefit():
  try:
    cursor = conn.cursor()
    q1 = "SELECT id_benefit FROM benefitwisata ORDER BY id_benefit DESC LIMIT 1"
    cursor.execute(q1)

    id_awal = cursor.fetchone() #ini haslnya bentuk array. di pecahkan dlu

    if len(id_awal) == 0:
      keyword = "BF001"
      return keyword
    else:
      idAwalToStr = id_awal[0]
      substrIdAwal = idAwalToStr[2:5] #ambil str di index 1, sampai ke "karakter ke 5"
      toIntId = int(substrIdAwal) + 1
      keyword = "BF" + str(toIntId).zfill(3) #semacam padleft, dia nambahin zero didepan.
      return keyword
    
  except Exception as e:
    print(e)
  finally:
    cursor.close()

@app.post('/insertPaket')
async def isiPaket(
  request: Request,
  user : JwtAuthorizationCredentials = Security(access_security),
):
  try:
    cursor = conn.cursor()

    form_data = await request.form()

    nama_paket = form_data.get('nama_paket')
    subjudul_paket = form_data.get('subjudulpaket')
    harga_paket = form_data.get('harga_paket')
    id_bis = form_data.get('id_bis')
    id_rute = form_data.get('id_rute')
    tgl_brkt = form_data.get('tgl_brkt')
    tgl_balik = form_data.get('tgl_balik')
    jambrkt = form_data.get('jambrkt')
    jambalik = form_data.get('jambalik')
    jlhpenumpang = form_data.get('jlhpenumpang')
    gbrpaket: Optional[UploadFile] = form_data.get('gbrpaket')
    benefit = form_data.get('benefit')
    buatId = idPaket()

    arrBenefit = []
    curr_benefit = ""

    for char in benefit:
      if char != ",": #jika dlm str benefit g ad koma, += trus setiap huruf ke dlm curr_benefit
        curr_benefit += char
      else: #jika ad koma, append curr_benefit lalu distripkan, lalu reset curr_benefit="" 
        if curr_benefit:
          arrBenefit.append(curr_benefit.strip())
          
        curr_benefit = ""

    # print(curr_benefit) # kalo di array kan, += selalu ambil value terakhir
    # ini buat kondisi pertama kalo udh di paling akhir, ga ad inputan. ya di append
    arrBenefit.append(curr_benefit.strip()) # makany ini buat append str yg paling ujung.

    idBenefit = fnIdBenefit()

    for i in arrBenefit:
      try:
        qBenefit = "INSERT INTO benefitwisata(id_benefit, benefit) VALUES(%s, %s)"
        cursor.execute(qBenefit, (idBenefit, i))
        conn.commit()
      except ValueError as e:
        print(e)

    q1 = "SELECT * FROM rute WHERE id = %s"
    cursor.execute(q1, (id_rute, ))
    items = cursor.fetchone()
    print(items)

    filename = f"{uuid.uuid4()}.jpg"
    file_location = os.path.join("images/gbrpaket", filename)

    #read & save file
    content = await gbrpaket.read()
    with open(file_location, "wb") as f:
      f.write(content)

    q2 = """
      INSERT INTO paketwisata(
        id_paket, nama_paket, harga_paket, id_bis, 
        rute_awal, rute_akhir, tgl_brkt, tgl_balik, 
        gbrpaket, subjudulpaket, jambrkt, jambalik,
        jlhpenumpang, id_benefit
      ) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    cursor.execute(q2, 
      (buatId, nama_paket, harga_paket, id_bis, items[1], items[2], 
       tgl_brkt, tgl_balik, filename, subjudul_paket, jambrkt, jambalik, jlhpenumpang, idBenefit)
    )
    conn.commit()

  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  
  finally:
    cursor.close()

@app.put('/updatePaket/{id_paket}')
async def updatePaket(
  id_paket: str,
  request: Request,
  user : JwtAuthorizationCredentials = Security(access_security),
):
  try:
    cursor = conn.cursor()

    form_data = await request.form()

    nama_paket = form_data.get('nama_paket')
    subjudul_paket = form_data.get('subjudulpaket')
    harga_paket = form_data.get('harga_paket')
    id_bis = form_data.get('id_bis')
    id_rute = form_data.get('id_rute')
    tgl_brkt = form_data.get('tgl_brkt')
    tgl_balik = form_data.get('tgl_balik')
    jambrkt = form_data.get('jambrkt')
    jambalik = form_data.get('jambalik')
    jlhpenumpang = form_data.get('jlhpenumpang')
    gbrpaket: Optional[UploadFile] = form_data.get('gbrpaket')
    benefit = form_data.get('benefit')
    idBenefit = form_data.get('id_benefit')

    # Logic Benefit
    splitBenefit = benefit.split(",") # Ini Belum dalam bentuk list
    lstBenefit = list(splitBenefit)
    try:
      # Ini ga bisa pake query update. karena dia bkl kereplace semua
      # Step 1. Delete Dlu Data dengan ID yang mau di input
      delQuery = "DELETE FROM benefitwisata WHERE id_benefit = %s"
      cursor.execute(delQuery, (idBenefit, ))

      # Step 2. Input Benefit Baru
      insertQuery = "INSERT INTO benefitwisata (benefit, id_benefit) VALUES (%s, %s)"
      for i in lstBenefit:
        cursor.execute(insertQuery, (i.strip(), idBenefit))  # Use the current benefit
      conn.commit()  # Commit semua Insert

    except Exception as e:
      print(e)
      conn.rollback()
    # End Logic benefit

    if gbrpaket is not None:
      filename = f"{uuid.uuid4()}.jpg"
      file_location = os.path.join("images/gbrpaket", filename)

      #read & save file
      content = await gbrpaket.read()
      with open(file_location, "wb") as f:
        f.write(content)

      q1 = "SELECT * FROM rute WHERE id = %s"
      cursor.execute(q1, (id_rute, ))
      items = cursor.fetchone()

      q2 = """
        UPDATE paketwisata SET nama_paket = %s, harga_paket = %s, id_bis = %s, rute_awal = %s,
        rute_akhir = %s, tgl_brkt = %s, tgl_balik = %s, gbrpaket = %s, subjudulpaket = %s,
        jambrkt = %s, jambalik = %s, jlhpenumpang = %s, id_benefit = %s
        WHERE id_paket = %s
      """
      cursor.execute(q2, 
        (nama_paket, harga_paket, id_bis, items[1], items[2], 
        tgl_brkt, tgl_balik, filename, subjudul_paket, jambrkt, jambalik, jlhpenumpang, idBenefit, id_paket)
      )
      conn.commit()
    else:
      # Jika g ad gbr
      q1 = "SELECT * FROM rute WHERE id = %s"
      cursor.execute(q1, (id_rute, ))
      items = cursor.fetchone()

      q2 = """
        UPDATE paketwisata SET nama_paket = %s, harga_paket = %s, id_bis = %s, rute_awal = %s,
        rute_akhir = %s, tgl_brkt = %s, tgl_balik = %s, subjudulpaket = %s,
        jambrkt = %s, jambalik = %s, jlhpenumpang = %s, id_benefit = %s
        WHERE id_paket = %s
      """
      cursor.execute(q2, 
        (nama_paket, harga_paket, id_bis, items[1], items[2], 
        tgl_brkt, tgl_balik, subjudul_paket, jambrkt, jambalik, jlhpenumpang, idBenefit, id_paket)
      )
      conn.commit()


  except HTTPException as e:
    return JSONResponse(content={"Error": str(e)}, status_code=e.status_code)
  
  finally:
    cursor.close()

@app.delete('/paket/{id}')
async def delpaket(
  id: str,
  user: JwtAuthorizationCredentials = Security(access_security)
) :
  try:
    cursor = conn.cursor()
    # print(data)

    q1 = "DELETE FROM paketwisata WHERE id_paket = %s"
    cursor.execute(q1, (id, ))
    conn.commit()

    return JSONResponse(content={"Success": "Lala"}, status_code=200)

  except HTTPException as e:
    return JSONResponse(content={"ErrorWoi": str(e)}, status_code=e.status_code)
  finally:
    cursor.close()

@app.get('/export_trans')
async def exportExcel(
  tglawal: str = Query(...),
  tglakhir: str = Query(...)
):
  try:
    # step 1 fetch query spt biasa
    cursor = conn.cursor()
    query = """
      SELECT t.id_trans, t.tgl_trans, t.email_cust, t.id_staff, t.status_trans, 
      t.total_harga, t.metode_byr, t.user_bayar, t.kembalian, dt.tgl_pergi, 
      dt.tgl_balik, b.id_rute, pw.nama_paket FROM transaksi t 
      INNER JOIN "detailTransaksi" dt ON t.id_trans = dt.id_trans
      INNER JOIN bis b ON dt.id_bis = b.id_bis
      LEFT JOIN paketwisata pw ON t.id_paket = pw.id_paket
      WHERE t.tgl_trans BETWEEN %s AND %s
      AND (t.status_trans = 'CANCELLED' OR t.status_trans = 'COMPLETED') 
    """
    cursor.execute(query, (tglawal, tglakhir))
    items = serialize_data(cursor.fetchall())

    column_name = []

    for kol in cursor.description:
      column_name.append(kol[0])

    df = pd.DataFrame(items, columns=column_name)

    # Step 2. Buat Workbook Excel lalu tambahkan worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "LAP. PENJUALAN"

    # Step 3. Tambahkan Header kaya judul lalu di merge cell
    corporate_name = "BUS_HUB"
    worksheet.merge_cells('A1:M1') #merge a1 sampe M1
    corp_cell = worksheet['A1']
    corp_cell.value = corporate_name

    # Step 4. Center text corporate cell
    corp_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True, size=18)
    for cell in worksheet[1]: # index row excel start dr 1. ini ceritany mw tebalin bushub
      cell.font = header_font

    # Step 5. Tambah Keterangan dibawah nama Korporat yg header di A1
    corporate_ket = "LAPORAN PENJUALAN"
    worksheet.merge_cells('A2:M2') 
    ket_cell = worksheet['A2']
    ket_cell.value = corporate_ket

    # Step 6. Center corporate addr
    ket_cell.alignment = Alignment(horizontal='center', vertical='center')
    header2_font = Font(bold=True, size=15)
    for cell in worksheet[2]: #Tebalin row kedua
      cell.font = header2_font
    
    worksheet.append([]) #buat space kosong di rowsnya

    # Step 7. Tambahkan Header dari nama column_name
    worksheet.append(column_name)
    header3_font = Font(bold=True, size=11)
    for cell in worksheet[4]: #index ke 4. yg nama column
      cell.font = header3_font

    # Step 8. Tambahin data dataframe ke dalam worksheet
    for row in dataframe_to_rows(df, index=False, header=False):
      worksheet.append(row)
      
    # Step 9. Simpan Workbook ke file
    file_path = "datapenjualan_bushub.xlsx"
    workbook.save(file_path)

    # Step 10. Return Excelnya sbg FileResponse
    return FileResponse(
      file_path,
      media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      filename="datapenjualan_bushub.xlsx"
    )
    
  except Exception as e:
    print(e)

  finally:
    cursor.close()
